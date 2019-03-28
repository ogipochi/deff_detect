import mojimoji
import re
from difflib import SequenceMatcher
import numpy as np
import math
from openpyxl import Workbook
from openpyxl.styles import borders

class TextToList:
    def __init__(self,text,version,conversion_table,heroin="ヒロイン",similarity=1.0,name_detect=True):
        self.text = text
        self.version = int(version)
        self.conversion_table = conversion_table
        self.name_origin_list = []
        self.heroin = heroin
        self.default_txt_obj = {
            "type":"mono",
            "content":"",
            "remarks":"",
            "name":[]
        }
        self.name_detect = name_detect
        self.everyone_patterns = [
            "$全員","＄全員","$ぜんいん","＄ぜんいん","$ゼンイン",
            "＄ゼンイン","$皆","＄皆","$みんな","＄みんな","$ミンナ",
            "＄ミンナ",
        ]
        self.converted_text_obj_list = []
        self.name_eval_list = []  # 名前の可能性のあったもの
        self.name_eval = ""     # 名前の可能性のあるものは一旦ここに入れる
    def _split_with_line_break(self):
        text_list_interim = self.text.splitlines()  # このリストには無駄な改行が含まれている可能性がある
        self.text_list = []
        pattern = re.compile(r'\s+')
        for num,text in enumerate(text_list_interim):
            text = text.strip()
            self.text_list.append(text)
        
        return self.text_list
    def _convert_text_to_obj_list_v2(self):
        txt_obj = self.default_txt_obj.copy()
        name_container = []
        text_type = "mono"
        for num,text in enumerate(self.text_list):

            if len(text) == 0:
                
                txt_obj["type"] = "blank"
                txt_obj["content"] = ""
                txt_obj["remarks"] = ""
                txt_obj["name"] = []
                self.converted_text_obj_list.append(txt_obj)
                txt_obj = self.default_txt_obj.copy()
                
                continue
            
            if "【" in text and "】" in text:
                txt_obj["type"] = "addition"
                txt_obj["content"] = text
                txt_obj["remarks"] = ""
                txt_obj["name"] = []
                self.converted_text_obj_list.append(txt_obj)
                txt_obj = self.default_txt_obj.copy()

                name_container = []
                text_type = "mono"
                continue
            elif "SE" in text:
                txt_obj["type"] = "addition"
                txt_obj["content"] = text
                txt_obj["remarks"] = ""
                txt_obj["name"] = []
                self.converted_text_obj_list.append(txt_obj)
                txt_obj = self.default_txt_obj.copy()
                name_container = []
                text_type = "mono"
                continue
            elif text[0] in ["<","＜","〈"] and text[-1] in [">","＞","〉"]:
                txt_obj["type"] = "addition"
                txt_obj["content"] = text
                txt_obj["remarks"] = ""
                txt_obj["name"] = []
                self.converted_text_obj_list.append(txt_obj)
                txt_obj = self.default_txt_obj.copy()
                name_container = []
                text_type = "mono"
                continue
            elif text in ["T","Ｔ"]:
                text_type = "mono"
                name_container = []
                continue
            elif text in self.conversion_table:
                text_type = "dialog"
                name_container = [self.conversion_table[text]]
                continue
            elif text == self.heroin:
                name_container = [self.heroin]
                text_type = "hero"
                continue
            elif re.split("[)(）（]",text)[0] in self.conversion_table:
                text_type = "dialog"
                name_container = [self.conversion_table[re.split("[)(）（]",text)[0]]]
                txt_obj["remarks"] = re.split("[)(）（]",text)[-1]
                continue
            elif self.name_detect and len(name_container)==0 and len(list(text))<=6:
                self.name_eval_list.append(text)
                name_container = [text]
                text_type = "dialog"
                continue
            else:

                txt_obj["content"] = text
                txt_obj["type"] = text_type
                txt_obj["name"] = name_container
                self.converted_text_obj_list.append(txt_obj)
                txt_obj = self.default_txt_obj.copy()
                continue
    def _convert_text_to_obj_list_v1(self):
        txt_obj = self.default_txt_obj.copy()

        for num,text in enumerate(self.text_list):
            # 空行 ##############################################################
            if len(list(text))==0:
                txt_obj["type"] = "blank"
                txt_obj["content"] = ""
                txt_obj["remarks"] = ""
                txt_obj["name"] = []
                self.converted_text_obj_list.append(txt_obj.copy())
                if self.name_eval:
                    self.name_eval=""
                continue
            ##名前の可能性のあったもの########################################################
            if self.name_eval and self.name_detect:
                if list(text)[0] in ["「","『","｢"]:
                    txt_obj["type"] = "dialog"
                    txt_obj["name"] = [self.name_eval]
                    self.name_eval_list.append(self.name_eval)
                    txt_obj["content"] = text
                    self.converted_text_obj_list.append(txt_obj.copy())
                    self.name_eval = ""
                    continue
                else:
                    txt_obj["type"] = "mono"
                    txt_obj["content"] = self.name_eval
                    self.name_eval = ""
                    self.converted_text_obj_list.append(txt_obj.copy())
                    txt_obj = self.default_txt_obj.copy()
            if txt_obj["type"] == "dialog" or txt_obj["type"] == "hero":
                if list(text)[0] in ["「","｢","『"]:
                    if not (list(text)[-1] in ["」","｣","』"]) and "※" in text:
                        text_splitted = text.split("※")
                        txt_obj["content"] = text_splitted[0].strip()
                        txt_obj["remarks"] = "※" + text_splitted[-1].strip()
                        self.converted_text_obj_list.append(txt_obj.copy())
                        txt_obj["remarks"] = ""
                        continue
                    else:
                        txt_obj["content"] = text
                        
                        self.converted_text_obj_list.append(txt_obj.copy())
                        txt_obj["remarks"] = ""
                        continue

            if list(text)[0] in ["○","◯",";"]:
                txt_obj["type"] = "addition"
                txt_obj["content"] = text
                self.converted_text_obj_list.append(txt_obj.copy())
                continue
            elif "※" in text or "・" in text:
                text_splited = text.split("※")
                if len(text_splited) > 1:
                    text = text_splited[0]
                    txt_obj["remarks"] = "※" + text_splited[-1]
                text_splited = text.split("・")
                txt_obj["name"] = [text.strip() for text in text_splited]
                txt_obj["type"] = "dialog"
                for i,name in enumerate(txt_obj["name"]):
                    if name in self.conversion_table:
                        txt_obj["name"][i] = self.conversion_table[name]
                    else:
                        self.name_eval_list.append(name)
                continue
            elif text in self.conversion_table:
                txt_obj["name"] = [self.conversion_table[text]]
                txt_obj["type"] = "dialog"
                continue
            elif text == self.heroin:
                txt_obj["type"] = "hero"
                txt_obj["content"] = ""
                txt_obj["name"] = [self.heroin]
                continue
            elif len(list(text))<10 and self.name_detect:
                self.name_eval = text
                continue
            else:
                txt_obj["type"] = "mono"
                txt_obj["content"] = text
                self.converted_text_obj_list.append(txt_obj.copy())
                continue

    
    def __convert_text_to_obj_list_v2(self):
        txt_obj = self.default_txt_obj.copy()
        line_break_names = [] # 空の改行で会話文が区切られたときのために名前を記憶しておくセル
        for num,text in enumerate(self.text_list):
            if len(list(text))==0:
                # 空行の場合
                if len(txt_obj["content"]):
                    
                    # もしtxt_objが初期化されずに残っていたら
                    # ここまでのtxt_objは会話文なので追加しなければ
                    self.converted_text_obj_list.append(txt_obj.copy())
                    # contentとremarksだけ初期化
                    txt_obj["content"] = ""
                    txt_obj["remarks"] = ""
                    # 更に空の改行分もある
                    blank_obj = self.default_txt_obj.copy()
                    blank_obj["type"] = "blank"
                    blank_obj["content"] = ""
                    blank_obj["remarks"] = ""
                    blank_obj["name"] = ""
                    self.converted_text_obj_list.append(blank_obj.copy())
                    continue
                elif len(txt_obj["remarks"]) and len(txt_obj["content"])==0:
                    # もしtxt_objが初期化されずに残っていたら
                    # ここまでのtxt_objは会話文なので追加しなければ
                    txt_obj["name"] = ""
                    self.converted_text_obj_list.append(txt_obj.copy())
                    # contentとremarksだけ初期化
                    txt_obj["content"] = ""
                    txt_obj["remarks"] = ""
                    # 更に空の改行分もある
                    blank_obj = self.default_txt_obj.copy()
                    blank_obj["type"] = "blank"
                    blank_obj["content"] = ""
                    blank_obj["remarks"] = ""
                    blank_obj["name"] = ""
                    self.converted_text_obj_list.append(blank_obj.copy())
                    continue
                else:
                    txt_obj["type"] = "blank"
                    txt_obj["content"] = ""
                    txt_obj["remarks"] = ""
                    txt_obj["name"] = ""
                    self.converted_text_obj_list.append(txt_obj.copy())
                    txt_obj = self.default_txt_obj.copy()
                    # 名前の初期化
                    line_break_names = []
            elif text in ["//END","／／END","//ＥＮＤ"]:
                txt_obj["type"] = "option"
                txt_obj["content"] = ""
                txt_obj["remarks"] = "//END"
                txt_obj["name"] = ""
                self.converted_text_obj_list.append(txt_obj.copy())
                txt_obj = self.default_txt_obj.copy()
                # 名前の初期化
                line_break_names = []
            elif len(line_break_names)>0 and (list(text)[0] in ["/","／"] and list(text)[1] in ["/","／"]):
                txt_obj["remarks"] = text
                # このまま名前が出ずに空行が来た場合はmonoになるため
                # 入れておく
                txt_obj["type"] = "mono"
                txt_obj["name"]
                continue
            
            elif list(text)[0] in ["［","[","【"] and ("]" in list(text) or "］" in list(text) or "】" in list(text)):
                # 名前の初期化
                line_break_names = []
                # txt_objの初期化
                txt_obj = self.default_txt_obj.copy()
                name_list = re.split("[［］\[\]・]",text)
                for name in name_list:
                    name = name.strip()
                    if len(list(name))==0:
                        # からの場合次へ
                        continue
                    if list(name)==self.heroin:
                        line_break_names.append(self.conversion_table[name])
                    elif list(name)[0] in ["（","("] and list(name)[-1] in ["）",")"]:
                        txt_obj["remarks"] += name
                        continue
                    elif list(name)[0] in ["/","／","＼"] and list(name)[1] in ["/","／","＼"]:
                        txt_obj["remarks"] += name
                        continue
                    elif name in self.conversion_table:
                        line_break_names.append(self.conversion_table[name])
                    elif name in ["全員","ぜんいん","ゼンイン","皆","みんな","ミンナ"]:
                        line_break_names.append("$全員")
                        continue
                    else:
                        line_break_names.append(name)
                        continue
                txt_obj["name"] = line_break_names
            elif list(text)[0] in ["/","／","＼"] and list(text)[1] in ["/","／","＼"]:
                txt_obj = self.default_txt_obj.copy()
                txt_obj["type"] = "mono"
                txt_obj["content"] = ""
                txt_obj["remarks"] = text
                self.converted_text_obj_list.append(txt_obj.copy())
                # txt_objの初期化
                txt_obj = self.default_txt_obj.copy()
                # 名前の初期化
                line_break_names = []
            elif list(text)[0] in ["◯","○"]:
                txt_obj = self.default_txt_obj.copy()
                txt_obj["type"] = "mono"
                txt_obj["content"] = ""
                txt_obj["remarks"] = text
                self.converted_text_obj_list.append(txt_obj.copy())
                # txt_objの初期化
                txt_obj = self.default_txt_obj.copy()
                # 名前の初期化
                line_break_names = []
            
            else:
                if len(txt_obj["name"])==0:
                    txt_obj["remarks"] = text
                    self.converted_text_obj_list.append(txt_obj.copy())
                    # txt_objの初期化
                    txt_obj = self.default_txt_obj.copy()
                    continue
                elif txt_obj["name"][0] == self.heroin:
                    txt_obj["type"] = "hero"
                else:
                    txt_obj["type"] = "dialog"
                if len(txt_obj["content"])>0:
                    txt_obj["content"] += "\r\n"    
                txt_obj["content"] += text
                continue
        return self.converted_text_obj_list      
    def convert(self):
        self._split_with_line_break()
        if self.version==1:
            self._convert_text_to_obj_list_v1()
            return self.converted_text_obj_list
        elif self.version==2:
            self._convert_text_to_obj_list_v2()
            return self.converted_text_obj_list
        else:
            return None

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import PatternFill
import openpyxl
import time


class DeffDetecter:
    def __init__(self , data_frames , data_frame_ids , text_list , version=1, list_window=30 , initial_row=3 , similarity=0.6 , dialog_sheet_col=7):
        self.data_frames = data_frames
        self.data_frame_ids = data_frame_ids
        self.text_list = text_list
        self.version = version
        self.list_window = list_window
        self.similarity = similarity
        self.dialog_sheet_col = dialog_sheet_col
        # 値の初期化
        self.now_id = 0
        self.initial_row = initial_row
        self.deff_list = []

        # テンプレート
        self.default_deff_elem = {
            "sheet_name":"",
            "type":"del",   # del , alt , add , none
            "original_text":"",
            "original_name":"",
            "alt_text":"",
            "alt_name" : "",
            "row":"",
        }
        self.pre_convert_rule_list = [
            {
                "original":"<param=heroine_last_name>",
                "alt":"ヒロイン",
            },
            {
                "original":"<param=heroine_first_name>",
                "alt":"苗字",
            },
        ]
    def similar(self,a,b):
        return SequenceMatcher(None, a, b).ratio()
    def _detect_func_v1(self):
        for i in sorted(self.data_frame_ids):
            sheet_name = self.data_frame_ids[i]
            df = self.data_frames[sheet_name]
            for row_num,row_data in enumerate(df.iterrows()):
                # print("DEFF:",self.deff_list)
                text = row_data[1][self.dialog_sheet_col]
                if type(text) == type(0.6):  # float削除用
                    continue
                else:
                    text_normalized = text
                    for pre_convert_rule in  self.pre_convert_rule_list:
                        text_normalized = text_normalized.replace(pre_convert_rule["original"],pre_convert_rule["alt"])
                    deff_elem = self.default_deff_elem.copy()
                    deff_elem["sheet_name"] = sheet_name
                    deff_elem["row"] = int(row_num) + (self.initial_row - 1)
                    deff_elem["original_text"] = text
                    
                    look_up_text_list = self.text_list[0:self.list_window]
                    for look_up_id,look_up_text in enumerate(look_up_text_list):
                        
                        find_token = False
                        if self.similar(look_up_text["content"],text_normalized) == 1:
                            find_token = True
                            deff_elem["type"] = "none"
                            if not look_up_id==0:
                                for add_id in range(look_up_id):
                                    add_text = look_up_text_list[add_id]
                                    if add_text["type"] in ["blank","addition"]:
                                        continue
                                    addition_elem = self.default_deff_elem.copy()
                                    addition_elem["type"] = "add"
                                    addition_elem["alt_text"] = look_up_text_list[add_id]["content"]
                                    addition_elem["alt_name"] = look_up_text_list[add_id]["name"]
                                    self.deff_list.append(addition_elem)
                            self.deff_list.append(deff_elem)
                            self.text_list = self.text_list[(look_up_id+1):]
                            
                            break
                        elif self.similar(look_up_text["content"],text_normalized) >= self.similarity:
                            find_token = True
                            deff_elem["type"] = "alt"
                            deff_elem["alt_text"] = look_up_text["content"]
                            if not look_up_id==0:
                                for add_id in range(look_up_id):
                                    add_text = look_up_text_list[add_id]
                                    if add_text["type"] in ["blank","addition"]:
                                        continue
                                    addition_elem = self.default_deff_elem.copy()
                                    addition_elem["type"] = "add"
                                    addition_elem["alt_text"] = add_text["content"]
                                    addition_elem["alt_name"] =add_text["name"]
                                    self.deff_list.append(addition_elem)
                                # self.deff_list.append(deff_elem)
                                # self.text_list = self.text_list[(look_up_id+1):]
                            self.deff_list.append(deff_elem)
                            self.text_list = self.text_list[(look_up_id+1):]
                            break
                    if not find_token:
                        deff_elem["type"] = "del"
                        self.deff_list.append(deff_elem)
        for add_id,add_text in enumerate(self.text_list):
            if add_text["type"] in ["blank","addition"]:
                continue
            addition_elem = self.default_deff_elem.copy()
            addition_elem["type"] = "add"
            addition_elem["alt_text"] = add_text["content"]
            addition_elem["alt_name"] = add_text["name"]
            self.deff_list.append(addition_elem)
    def _paste_text_list_together(self):
        text_list_pasted = []
        token = False
        text_stock = None
        for text in self.text_list:
            if text["type"] =="blank":
                token = False
                if text_stock:
                    text_list_pasted.append(text_stock)
            elif token and text_stock["type"] == text["type"] and text_stock["name"] == text["name"]:
                text_stock["content"] += "\n"
                text_stock["content"] += text["content"]
            else:
                token = True
                text_stock = text
            
        self.text_list  =text_list_pasted
    def _detect_func_v2(self):
        self._paste_text_list_together()
        for i in sorted(self.data_frame_ids):
            sheet_name = self.data_frame_ids[i]
            print(sheet_name)
            df = self.data_frames[sheet_name]
            for row_num , row_data in enumerate(df.iterrows()):
                text = row_data[1][6]
                text_type = row_data[1][2]
                if not text_type in ["吹き出し","ポエム"]:
                    continue
                if type(text) == type(0.6):
                    continue
                text_normalized = re.sub(r"<color=#[a-z0-9]{6}>","",text)
                text_normalized = text_normalized.replace("</color>","")
                text_normalized = text_normalized.replace("＊名前＊","ヒロイン")
                deff_elem = self.default_deff_elem.copy()
                deff_elem["sheet_name"] = sheet_name
                deff_elem["row"] = int(row_num) + (self.initial_row - 1)
                deff_elem["original_text"] = text
                look_up_text_list = self.text_list[0:self.list_window]
                if text == "……なるほどね":
                    log = True
                else:
                    log = False
                if len(text_normalized) == 0:
                    deff_elem["type"] = "blank"
                    self.deff_list.append(deff_elem)
                    continue
                for look_up_id,look_up_text in enumerate(look_up_text_list):
                    # 改行を削除しておく
                    #look_up_text_normalized = look_up_text["content"].replace("\r\n","")
                    #look_up_text_normalized = look_up_text_normalized.replace("\n","")
                    #look_up_text_normalized = look_up_text_normalized.replace("\r","")
                    look_up_text_normalized = look_up_text["content"]
                    #text_normalized = text_normalized.replace("\r\n","")
                    #text_normalized = text_normalized.replace("\n","")
                    #text_normalized = text_normalized.replace("\r","")
                    if log:
                        print(look_up_text_normalized)
                    if self.similar(look_up_text_normalized,text_normalized) == 1:
                        if log:
                            print("SIMILAR")
                        find_token = True
                        # look_upの中の追加要素を探る
                        deff_elem["type"] = "none"
                        deff_elem["alt_text"] = look_up_text["content"]
                        if not look_up_id==0:
                            for add_id in range(1,look_up_id):
                                add_text = look_up_text_list[add_id]
                                if add_text["type"] in ["blank","addition"]:
                                    continue
                                addition_elem = self.default_deff_elem.copy()
                                addition_elem["type"] = "add"
                                addition_elem["alt_text"] = add_text["content"]
                                addition_elem["alt_name"] = add_text["name"]
                                self.deff_list.append(addition_elem)
                                
                        self.deff_list.append(deff_elem)
                        self.text_list = self.text_list[(look_up_id+1):]
                        
                        break
                    elif len(text_normalized) > 4 and self.similar(look_up_text_normalized,text_normalized) >= self.similarity:
                        if log:
                            print("OO")
                        find_token = True
                        deff_elem["type"] = "alt"
                        deff_elem["alt_text"] = look_up_text["content"]
                        if not look_up_id==0:
                            for add_id in range(1,look_up_id):
                                add_text = look_up_text_list[add_id]
                                if add_text["type"] in ["blank","addition"]:
                                    continue
                                addition_elem = self.default_deff_elem.copy()
                                addition_elem["type"] = "add"
                                addition_elem["alt_text"] = add_text["content"]
                                addition_elem["alt_name"] =add_text["name"]
                                self.deff_list.append(addition_elem)
                                
                        self.deff_list.append(deff_elem)
                        self.text_list = self.text_list[(look_up_id+1):]
                        break
                    elif look_up_id>self.list_window:
                        find_token = False
                        break
                    else:
                        continue
                if not find_token:
                    deff_elem["type"] = "del"
                    self.deff_list.append(deff_elem)
        for add_id,add_text in enumerate(self.text_list):
            if add_text["type"] in ["blank","addition"]:
                continue
            addition_elem = self.default_deff_elem.copy()
            addition_elem["type"] = "add"
            addition_elem["alt_text"] = add_text["content"]
            addition_elem["alt_name"] = add_text["name"]
            self.deff_list.append(addition_elem)
    def detect(self):
        if self.version==1:
            self._detect_func_v1()
        elif self.version==2:
            self._detect_func_v2()
        else:
            return False
        return self.deff_list
    def create_wb(self,waiting_id):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'sheet名'
        ws.column_dimensions['A'].width = 15
        ws['B1'] = 'row'
        ws['C1'] = 'キャラ名(元)'
        ws['D1'] = '元エクセルテキスト'
        ws.column_dimensions['D'].width = 70
        ws['E1'] = 'キャラ名(変更先)'
        ws['F1'] = '変更部分'
        ws.column_dimensions['F'].width = 70
        ws['G1'] = 'タイプ'
        print("this")
        for i,deff_elem in enumerate(self.deff_list):
            line_height = 1
            if deff_elem["original_text"].count("\n") > deff_elem["alt_text"].count("\n"):
               line_height = deff_elem["original_text"].count("\n")
            else:
               line_height = deff_elem["alt_text"].count("\n")
            if deff_elem["type"] == "del":
                color = "ff8a80"
            elif deff_elem["type"] == "add":
                color = "82b1ff"
            elif deff_elem["type"] == "alt":
                color = "b9f6ca"
            else:
                color = "FFFFFF"
            border = borders.Border(top=borders.Side(style=borders.BORDER_HAIR, color='eeeeee'),
                                 right=borders.Side(style=borders.BORDER_HAIR, color='eeeeee'),
                                 left=borders.Side(style=borders.BORDER_HAIR, color='eeeeee'),
                                 bottom=borders.Side(style=borders.BORDER_HAIR, color='eeeeee')
                                 )
            fill = openpyxl.styles.PatternFill(patternType='solid',
                                   fgColor=color)
            ws.row_dimensions[i+3].height = line_height * 35
            ws['A{}'.format(i+3)] = deff_elem["sheet_name"]
            ws['A{}'.format(i+3)].fill = fill
            ws['A{}'.format(i+3)].border = border

            ws['B{}'.format(i+3)] = deff_elem["row"]
            ws['B{}'.format(i+3)].fill = fill
            ws['B{}'.format(i+3)].border = border
            
            ws['C{}'.format(i+3)].fill = fill
            ws['C{}'.format(i+3)].border = border

            ws['D{}'.format(i+3)] = deff_elem["original_text"]
            ws['D{}'.format(i+3)].fill = fill
            ws['D{}'.format(i+3)].border = border

            ws['E{}'.format(i+3)] = ",".join(deff_elem["alt_name"])
            ws['E{}'.format(i+3)].fill = fill
            ws['E{}'.format(i+3)].border = border

            ws['F{}'.format(i+3)] = deff_elem["alt_text"]
            ws['F{}'.format(i+3)].fill = fill
            ws['F{}'.format(i+3)].border = border

            ws['G{}'.format(i+3)] = deff_elem["type"]
            ws['G{}'.format(i+3)].fill = fill
            ws['G{}'.format(i+3)].border = border

        wb.save("./converted_file/{}.xlsx".format(waiting_id))
        return "converted_file/{}.xlsx".format(waiting_id)
